from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors, Font
import locale, time


class Excel(object):
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path  # 目的在于在类的不同方法中共享此变量
        self.wb = load_workbook(excel_file_path)  # 表格对象
        self.ws = self.wb[self.wb.sheetnames[0]]  # 表格的第一个sheet对象
        # 获取第一个sheet不能用self.wb.ative，因为这个选择的是默认被打开的那个，如果表格在关闭前是打开的第三个sheet，那么下一次打开的时候就是默认在第三个sheet的
        # print(self.ws.title)

    def get_all_sheet_names(self):
        return self.wb.sheetnames

    def get_sheet_name_by_index(self, index):
        return self.wb.sheetnames[index - 1]

    def get_excel_file_path(self):  # 获取excel表格的路径
        return self.excel_file_path

    def create_sheet1(self, sheet_name, position=None):
        try:
            if position:
                self.wb.create_sheet(sheet_name, position)
            else:
                self.wb.create_sheet(sheet_name)
            self.save()  # 向表格里写入东西后一定要保存
            return True
        except Exception as e:
            print(e)
            return False

    def set_sheet_by_name(self, sheet_name):
        if sheet_name not in self.wb.sheetnames:
            print('%s sheet不存在，请重新设置！' % sheet_name)
            return False
        self.ws = self.wb[sheet_name]
        return True

    def set_sheet_by_index(self, index):
        self.ws = self.wb[self.get_sheet_name_by_index(index)]
        print('设定的sheet名称是:', self.ws.title)

    # 读值
    def get_cell_value(self, row_no, col_no, sheet_name=None):
        if sheet_name is not None:  # 参数设置了新的sheet
            result = self.set_sheet_by_name(sheet_name)
            if result == False:
                return None
        return self.ws.cell(row_no, col_no).value

    # 读某一行的值
    def get_row_values(self, row_no, sheet_name=None):
        cell_values = []  # 因为会返回多个值
        if sheet_name is not None:  # 参数设置了新的sheet
            result = self.set_sheet_by_name(sheet_name)
            if result == False:
                return None
        for cell in list(self.ws.rows)[row_no - 1]:
            cell_values.append(cell.value)
        return cell_values

    #读取某个sheet的所有行中的单元格内容，使用2维的列表进行存储
    def get_row_values(self,sheet_name=None):
        all_cell_values = []  # 所有的单元格的值均存入列表
        if sheet_name is not None:  # 参数设置了新的sheet
            result = self.set_sheet_by_name(sheet_name)
            if result == False:
                return None
        for row in list(self.ws.rows):
            row_values = []
            for cell in row:
                row_values.append(cell.value)
            all_cell_values.append(row_values)
        return all_cell_values

    # 读某一列的值
    def get_col_values(self, col_no, sheet_name=None):
        cell_values = []  # 因为会返回多个值
        if sheet_name is not None:  # 参数设置了新的sheet
            result = self.set_sheet_by_name(sheet_name)
            if result == False:
                return None
        for cell in list(self.ws.columns)[col_no - 1]:
            cell_values.append(cell.value)
        return cell_values

    # 读某个范围的值
    def get_some_values(self, min_row_no, min_col_no, max_row_no, max_col_no, sheet_name=None):
        if sheet_name is not None:  # 参数设置了新的sheet
            result = self.set_sheet_by_name(sheet_name)
            if result == False:
                return None
        values = []
        for i in range(min_row_no, max_row_no + 1):
            row_values = []  # 用来存一行的值
            for j in range(min_col_no, max_col_no + 1):
                row_values.append(self.ws.cell(row=i, column=j).value)
            values.append(row_values)  # 将读到的这一行值加入结果中
        return values

    # 保存
    def save(self):
        self.wb.save(self.excel_file_path)

    # 写
    def write_cell_value(self, row_no, col_no, value, style=None, sheet_name=None):
        if sheet_name is not None:  # 参数设置了新的sheet
            result = self.set_sheet_by_name(sheet_name)
            if result == False:
                return None
        if style is None:
            style = colors.BLACK
        elif style == 'red':
            style = colors.RED
        elif style == 'green':
            style = colors.GREEN
        self.ws.cell(row=row_no, column=col_no).font = Font(color=style)
        self.ws.cell(row=row_no, column=col_no, value=value)
        self.save()
        return True

    # 写时间
    def write_current_time(self, row_no, col_no, style=None, sheet_name=None):
        if sheet_name is not None:  # 参数设置了新的sheet
            result = self.set_sheet_by_name(sheet_name)
            if result == False:
                return None
        if style is None:
            style = colors.BLACK
        elif style == 'red':
            style = colors.RED
        elif style == 'green':
            style = colors.GREEN
        locale.setlocale(locale.LC_ALL, 'en')
        locale.setlocale(locale.LC_CTYPE, 'chinese')
        self.ws.cell(row=row_no, column=col_no).font = Font(color=style)
        self.ws.cell(row=row_no, column=col_no, value=time.strftime('%Y年%m月%d日 %H时%M分%S秒'))
        self.save()
        return True


if __name__ == '__main__':
    excel = Excel('c:\\sample.xlsx')
    # print(excel.get_excel_file_path())
    # print(excel.get_cell_value(1,1))
    # print(excel.get_cell_value(3,3))
    excel.set_sheet_by_name('xxx')  # 不存在的sheet
    excel.set_sheet_by_name('Sheet1')  # 存在的sheet
    # print(excel.get_cell_value(3,3))
    # print(excel.get_cell_value(3,3,'xxx')) #不存在的sheet读值时返回None
    # print(excel.get_row_values(1))
    # print(excel.get_row_values(1,'Sheet1'))
    # print(excel.get_row_values(1,'Sheet2'))
    # print(excel.get_col_values(1))
    # print(excel.get_col_values(1,'Sheet1'))
    # print(excel.get_col_values(1,'Sheet2'))
    # print(excel.get_some_values(1,1,4,4))
    # print(excel.get_some_values(1,1,3,3,'Sheet1'))
    # print(excel.get_some_values(1,1,3,3,'Sheet2'))
    # print(excel.write_cell_value(6,6,'你好'))
    # print(excel.write_cell_value(6,6,'你好','red'))
    # print(excel.write_current_time(7,7,'red'))
    # print(excel.get_all_sheet_names())
    # print(excel.get_all_sheet_name_by_index(1))
    # excel.set_sheet_by_index(2)
    print(excel.create_sheet1('uu1'))
